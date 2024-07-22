from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Protection
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.image import Image as Img


wb = Workbook()
sheet = wb.active

font0 = Font("Times", bold=False)
font1 = Font("Times", size='13', bold=True, color='00FF0000')
font2 = Font("Times", bold=True)
font3 = Font("Times", size='10', bold=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# ~ sheet.protection.sheet = True
# ~ sheet.protection.enable()

# ~ sheet["B8"].protection = Protection(locked=False, hidden=False)

for i in range(12,60):
	sheet["B"+str(i)].font = font0
	sheet["D"+str(i)].font = font0

img = Img("/home/pi/Spotcheck/logo.png")
img.height = 120
img.width = 600
img.anchor = 'B2'
sheet.add_image(img)

# ~ img = Img(self.base_window.qualitative_analysis_0.result_folder_path + "/result_capture.jpg")
# ~ img.anchor = 'H11'
# ~ sheet.add_image(img)

# ~ sheet["C10"] = self.base_window.qualitative_analysis_0.template_name

sheet.merge_cells(start_row=8, start_column=2, end_row=9, end_column=8)
sheet["B8"] = 'PHIẾU TRẢ KẾT QUẢ'
sheet["B8"].font = font1
sheet.cell(row=8,column=2).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)

sheet["B10"] = 'THÔNG TIN KHÁCH HÀNG'
sheet["B10"].font = font3
sheet["B11"] = 'Tên khách hàng: '
sheet["B11"].font = font0
sheet["B12"] = 'Địa chỉ: '
sheet["B12"].font = font0
sheet["B13"] = 'Số điện thoại: '
sheet["B13"].font = font0

sheet["B14"] = 'QUẢN LÝ DỮ LIỆU'
sheet["B14"].font = font3
sheet["B15"] = 'Test name: ' + 'test_name'
sheet["B15"].font = font0
sheet["B16"] = 'Test name: ' + 'test_name'
sheet["B16"].font = font0
sheet["B17"] = 'Date: ' + 'date'
sheet["B17"].font = font0

for r in range(19,68):
	for c in range(2,9):
		sheet.cell(row=r,column=c).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
		sheet.cell(row=r,column=c).border = thin_border

sheet.column_dimensions['B'].width = 26
sheet.column_dimensions['C'].width = 12
sheet.column_dimensions['D'].width = 12
sheet.column_dimensions['E'].width = 12
sheet.column_dimensions['F'].width = 12
sheet.column_dimensions['G'].width = 12
sheet.column_dimensions['H'].width = 12
sheet.row_dimensions[19].height = 40

sheet['B19'] = 'Tên mẫu'
sheet["B19"].font = font2
sheet["B19"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
sheet['C19'] = 'Vị trí'
sheet["C19"].font = font2
sheet["C19"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
sheet['D19'] = 'Giai đoạn'
sheet["D19"].font = font2
sheet["D19"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
sheet['E19'] = 'Bệnh'
sheet["E19"].font = font2
sheet["E19"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
sheet['F19'] = 'KQ Spotcheck'
sheet["F19"].font = font2
sheet["F19"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
sheet['G19'] = 'KQ điện di'
sheet["G19"].font = font2
sheet["G19"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
sheet['H19'] = 'Kết luận'
sheet["H19"].font = font2
sheet["H19"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')

wb.save('test.xlsx')
